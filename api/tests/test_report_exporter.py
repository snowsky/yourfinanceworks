"""
Unit tests for report export handlers
"""

import pytest
import io
import csv
from datetime import datetime
from unittest.mock import Mock, patch
from openpyxl import load_workbook

from core.services.report_exporter import (
    PDFExporter, CSVExporter, ExcelExporter, ReportExportService, ExportError
)
from core.schemas.report import (
    ReportData, ReportSummary, ReportMetadata, ReportType, ExportFormat
)


@pytest.fixture
def sample_report_data():
    """Create sample report data for testing"""
    return ReportData(
        report_type=ReportType.INVOICE,
        summary=ReportSummary(
            total_records=3,
            total_amount=1500.00,
            currency="USD",
            date_range={
                "start": datetime(2024, 1, 1),
                "end": datetime(2024, 1, 31)
            },
            key_metrics={
                "average_amount": 500.00,
                "paid_invoices": 2,
                "pending_invoices": 1
            }
        ),
        data=[
            {
                "invoice_number": "INV-001",
                "client_name": "Client A",
                "amount": 500.00,
                "status": "paid",
                "date": datetime(2024, 1, 15)
            },
            {
                "invoice_number": "INV-002", 
                "client_name": "Client B",
                "amount": 750.00,
                "status": "paid",
                "date": datetime(2024, 1, 20)
            },
            {
                "invoice_number": "INV-003",
                "client_name": "Client C", 
                "amount": 250.00,
                "status": "pending",
                "date": datetime(2024, 1, 25)
            }
        ],
        metadata=ReportMetadata(
            generated_at=datetime(2024, 1, 31, 10, 30, 0),
            generated_by=1,
            export_format=ExportFormat.JSON,
            generation_time=2.5
        ),
        filters={"date_from": "2024-01-01", "date_to": "2024-01-31"}
    )


@pytest.fixture
def empty_report_data():
    """Create empty report data for testing error cases"""
    return ReportData(
        report_type=ReportType.CLIENT,
        summary=ReportSummary(
            total_records=0,
            total_amount=0.00,
            currency="USD",
            key_metrics={}
        ),
        data=[],
        metadata=ReportMetadata(
            generated_at=datetime.now(),
            generated_by=1,
            export_format=ExportFormat.JSON
        ),
        filters={}
    )


@pytest.fixture
def company_data():
    """Sample company data for branding"""
    return {
        "name": "Test Company Inc.",
        "address": "123 Test Street, Test City, TC 12345",
        "phone": "+1-555-123-4567",
        "email": "info@testcompany.com"
    }


class TestPDFExporter:
    """Test cases for PDF export functionality"""
    
    def test_pdf_exporter_initialization(self, company_data):
        """Test PDF exporter initialization"""
        exporter = PDFExporter(company_data)
        assert exporter.company_data == company_data
        assert exporter.styles is not None
        assert 'ReportTitle' in exporter.styles.byName
        assert 'CompanyName' in exporter.styles.byName
    
    def test_pdf_export_success(self, sample_report_data, company_data):
        """Test successful PDF export"""
        exporter = PDFExporter(company_data)
        result = exporter.export_report(sample_report_data)
        
        assert isinstance(result, bytes)
        assert len(result) > 0
        # Check PDF header
        assert result.startswith(b'%PDF-')
    
    def test_pdf_export_without_company_data(self, sample_report_data):
        """Test PDF export without company data"""
        exporter = PDFExporter()
        result = exporter.export_report(sample_report_data)
        
        assert isinstance(result, bytes)
        assert len(result) > 0
        assert result.startswith(b'%PDF-')
    
    def test_pdf_export_with_empty_data(self, empty_report_data):
        """Test PDF export with empty data"""
        exporter = PDFExporter()
        
        # Should raise ExportError for empty data
        with pytest.raises(ExportError, match="Report data cannot be empty"):
            exporter.export_report(empty_report_data)
    
    def test_pdf_export_invalid_data(self):
        """Test PDF export with invalid data"""
        exporter = PDFExporter()
        
        with pytest.raises(ExportError, match="Report data is required"):
            exporter.export_report(None)
    
    def test_pdf_date_range_formatting(self, sample_report_data):
        """Test date range formatting in PDF"""
        exporter = PDFExporter()
        
        # Test with date range
        date_range = {
            "start": datetime(2024, 1, 1),
            "end": datetime(2024, 1, 31)
        }
        result = exporter._format_date_range(date_range)
        assert result == "2024-01-01 to 2024-01-31"
        
        # Test with no date range
        result = exporter._format_date_range(None)
        assert result == "All time"
        
        # Test with only start date
        date_range = {"start": datetime(2024, 1, 1)}
        result = exporter._format_date_range(date_range)
        assert result == "From 2024-01-01"


class TestCSVExporter:
    """Test cases for CSV export functionality"""
    
    def test_csv_exporter_initialization(self):
        """Test CSV exporter initialization"""
        exporter = CSVExporter()
        assert exporter is not None
    
    def test_csv_export_success(self, sample_report_data):
        """Test successful CSV export"""
        exporter = CSVExporter()
        result = exporter.export_report(sample_report_data)
        
        assert isinstance(result, str)
        assert len(result) > 0
        
        # Parse CSV to verify structure
        csv_reader = csv.DictReader(io.StringIO(result))
        rows = list(csv_reader)
        
        assert len(rows) == 3
        assert 'invoice_number' in rows[0]
        assert 'client_name' in rows[0]
        assert 'amount' in rows[0]
        assert 'status' in rows[0]
        
        # Check first row data
        assert rows[0]['invoice_number'] == 'INV-001'
        assert rows[0]['client_name'] == 'Client A'
        assert rows[0]['amount'] == '500.0'
        assert rows[0]['status'] == 'paid'
    
    def test_csv_export_with_none_values(self, sample_report_data):
        """Test CSV export with None values"""
        # Add None values to test data
        sample_report_data.data[0]['optional_field'] = None
        
        exporter = CSVExporter()
        result = exporter.export_report(sample_report_data)
        
        csv_reader = csv.DictReader(io.StringIO(result))
        rows = list(csv_reader)
        
        # None values should be converted to empty strings
        assert rows[0]['optional_field'] == ''
    
    def test_csv_export_with_datetime_values(self, sample_report_data):
        """Test CSV export with datetime values"""
        exporter = CSVExporter()
        result = exporter.export_report(sample_report_data)
        
        csv_reader = csv.DictReader(io.StringIO(result))
        rows = list(csv_reader)
        
        # Datetime should be in ISO format
        assert '2024-01-15T00:00:00' in rows[0]['date']
    
    def test_csv_export_with_empty_data(self, empty_report_data):
        """Test CSV export with empty data"""
        exporter = CSVExporter()
        
        with pytest.raises(ExportError, match="Report data cannot be empty"):
            exporter.export_report(empty_report_data)
    
    def test_csv_export_invalid_data(self):
        """Test CSV export with invalid data"""
        exporter = CSVExporter()
        
        with pytest.raises(ExportError, match="Report data is required"):
            exporter.export_report(None)


class TestExcelExporter:
    """Test cases for Excel export functionality"""
    
    def test_excel_exporter_initialization(self, company_data):
        """Test Excel exporter initialization"""
        exporter = ExcelExporter(company_data)
        assert exporter.company_data == company_data
    
    def test_excel_export_success(self, sample_report_data, company_data):
        """Test successful Excel export"""
        exporter = ExcelExporter(company_data)
        result = exporter.export_report(sample_report_data)
        
        assert isinstance(result, bytes)
        assert len(result) > 0
        
        # Load workbook to verify structure
        wb = load_workbook(io.BytesIO(result))
        
        # Check sheets exist
        assert 'Summary' in wb.sheetnames
        assert 'Report Data' in wb.sheetnames
        
        # Check summary sheet
        summary_ws = wb['Summary']
        assert 'Invoice Report Summary' in str(summary_ws['A1'].value)
        assert 'Company:' in str(summary_ws['A3'].value)
        assert company_data['name'] in str(summary_ws['B3'].value)
        
        # Check data sheet
        data_ws = wb['Report Data']
        assert data_ws['A1'].value == 'invoice_number'
        assert data_ws['B1'].value == 'client_name'
        assert data_ws['A2'].value == 'INV-001'
        assert data_ws['B2'].value == 'Client A'
    
    def test_excel_export_without_company_data(self, sample_report_data):
        """Test Excel export without company data"""
        exporter = ExcelExporter()
        result = exporter.export_report(sample_report_data)
        
        assert isinstance(result, bytes)
        assert len(result) > 0
        
        # Load workbook to verify it still works
        wb = load_workbook(io.BytesIO(result))
        assert 'Summary' in wb.sheetnames
        assert 'Report Data' in wb.sheetnames
    
    def test_excel_export_with_empty_data(self, empty_report_data):
        """Test Excel export with empty data"""
        exporter = ExcelExporter()
        
        with pytest.raises(ExportError, match="Report data cannot be empty"):
            exporter.export_report(empty_report_data)
    
    def test_excel_export_invalid_data(self):
        """Test Excel export with invalid data"""
        exporter = ExcelExporter()
        
        with pytest.raises(ExportError, match="Report data is required"):
            exporter.export_report(None)
    
    def test_excel_date_range_formatting(self):
        """Test date range formatting in Excel"""
        exporter = ExcelExporter()
        
        # Test with date range
        date_range = {
            "start": datetime(2024, 1, 1),
            "end": datetime(2024, 1, 31)
        }
        result = exporter._format_date_range_excel(date_range)
        assert result == "2024-01-01 to 2024-01-31"
        
        # Test with no date range
        result = exporter._format_date_range_excel(None)
        assert result == "All time"


class TestReportExportService:
    """Test cases for the main export service"""
    
    def test_service_initialization(self, company_data):
        """Test service initialization"""
        service = ReportExportService(company_data)
        assert service.company_data == company_data
        assert len(service.exporters) == 3
        assert ExportFormat.PDF in service.exporters
        assert ExportFormat.CSV in service.exporters
        assert ExportFormat.EXCEL in service.exporters
    
    def test_export_pdf(self, sample_report_data, company_data):
        """Test PDF export through service"""
        service = ReportExportService(company_data)
        result = service.export_report(sample_report_data, ExportFormat.PDF)
        
        assert isinstance(result, bytes)
        assert result.startswith(b'%PDF-')
    
    def test_export_csv(self, sample_report_data):
        """Test CSV export through service"""
        service = ReportExportService()
        result = service.export_report(sample_report_data, ExportFormat.CSV)
        
        assert isinstance(result, str)
        assert 'invoice_number' in result
        assert 'INV-001' in result
    
    def test_export_excel(self, sample_report_data, company_data):
        """Test Excel export through service"""
        service = ReportExportService(company_data)
        result = service.export_report(sample_report_data, ExportFormat.EXCEL)
        
        assert isinstance(result, bytes)
        
        # Verify it's a valid Excel file
        wb = load_workbook(io.BytesIO(result))
        assert 'Summary' in wb.sheetnames
        assert 'Report Data' in wb.sheetnames
    
    def test_unsupported_format(self, sample_report_data):
        """Test export with unsupported format"""
        service = ReportExportService()
        
        # Create a mock unsupported format
        with pytest.raises(ExportError, match="Unsupported export format"):
            service.export_report(sample_report_data, "unsupported")
    
    def test_get_supported_formats(self):
        """Test getting supported formats"""
        service = ReportExportService()
        formats = service.get_supported_formats()
        
        assert len(formats) == 3
        assert ExportFormat.PDF in formats
        assert ExportFormat.CSV in formats
        assert ExportFormat.EXCEL in formats
    
    def test_validate_export_format_valid(self):
        """Test export format validation with valid formats"""
        service = ReportExportService()
        
        assert service.validate_export_format("pdf") == ExportFormat.PDF
        assert service.validate_export_format("csv") == ExportFormat.CSV
        assert service.validate_export_format("excel") == ExportFormat.EXCEL
        assert service.validate_export_format("PDF") == ExportFormat.PDF
    
    def test_validate_export_format_invalid(self):
        """Test export format validation with invalid format"""
        service = ReportExportService()
        
        with pytest.raises(ExportError, match="Invalid export format"):
            service.validate_export_format("invalid")
    
    @patch('core.services.report_exporter.PDFExporter.export_report')
    def test_export_error_handling(self, mock_pdf_export, sample_report_data):
        """Test error handling during export"""
        # Mock PDF exporter to raise an exception
        mock_pdf_export.side_effect = Exception("PDF generation failed")
        
        service = ReportExportService()
        
        with pytest.raises(ExportError, match="Export failed"):
            service.export_report(sample_report_data, ExportFormat.PDF)


class TestExportErrorHandling:
    """Test cases for export error handling"""
    
    def test_export_error_creation(self):
        """Test ExportError creation"""
        error = ExportError("Test error message")
        assert str(error) == "Test error message"
    
    def test_base_exporter_not_implemented(self, sample_report_data):
        """Test base exporter raises NotImplementedError"""
        from core.services.report_exporter import BaseExporter
        
        exporter = BaseExporter()
        
        with pytest.raises(NotImplementedError):
            exporter.export_report(sample_report_data)
    
    def test_validation_with_missing_summary(self, sample_report_data):
        """Test validation with missing summary"""
        from core.services.report_exporter import BaseExporter
        
        sample_report_data.summary = None
        exporter = BaseExporter()
        
        with pytest.raises(ExportError, match="Report summary is required"):
            exporter.validate_report_data(sample_report_data)


class TestExportFormatting:
    """Test cases for export formatting and data handling"""
    
    def test_pdf_numeric_formatting(self, sample_report_data):
        """Test numeric formatting in PDF export"""
        # Add various numeric types to test data
        sample_report_data.data[0]['integer_value'] = 1000
        sample_report_data.data[0]['float_value'] = 1234.56
        sample_report_data.data[0]['large_number'] = 1000000.99
        
        exporter = PDFExporter()
        result = exporter.export_report(sample_report_data)
        
        assert isinstance(result, bytes)
        assert len(result) > 0
    
    def test_csv_special_characters(self, sample_report_data):
        """Test CSV export with special characters"""
        # Add special characters to test data
        sample_report_data.data[0]['description'] = 'Test with "quotes" and, commas'
        sample_report_data.data[0]['notes'] = 'Line 1\nLine 2'
        
        exporter = CSVExporter()
        result = exporter.export_report(sample_report_data)
        
        # Should handle special characters properly - CSV escapes quotes by doubling them
        assert '"Test with ""quotes"" and, commas"' in result
        csv_reader = csv.DictReader(io.StringIO(result))
        rows = list(csv_reader)
        assert rows[0]['description'] == 'Test with "quotes" and, commas'
    
    def test_excel_data_types(self, sample_report_data):
        """Test Excel export with various data types"""
        # Add various data types
        sample_report_data.data[0]['integer_value'] = 1000
        sample_report_data.data[0]['float_value'] = 1234.56
        sample_report_data.data[0]['boolean_value'] = True
        sample_report_data.data[0]['none_value'] = None
        
        exporter = ExcelExporter()
        result = exporter.export_report(sample_report_data)
        
        wb = load_workbook(io.BytesIO(result))
        data_ws = wb['Report Data']
        
        # Check that data types are preserved
        assert isinstance(data_ws['F2'].value, int)  # integer_value
        assert isinstance(data_ws['G2'].value, float)  # float_value
        assert data_ws['I2'].value is None or data_ws['I2'].value == ''  # none_value


if __name__ == "__main__":
    pytest.main([__file__])