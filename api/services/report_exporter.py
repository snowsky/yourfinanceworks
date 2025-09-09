"""
Report Export Handlers

This module provides export functionality for reports in multiple formats:
- PDF with professional formatting and company branding
- CSV for data analysis compatibility  
- Excel with multi-sheet support and formatting
"""

import csv
import io
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional, Union
from pathlib import Path

# PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Local imports
from schemas.report import ReportData, ExportFormat

logger = logging.getLogger(__name__)


class ExportError(Exception):
    """Custom exception for export-related errors"""
    pass


class BaseExporter:
    """Base class for all export handlers"""
    
    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def export_report(self, report_data: ReportData, **kwargs) -> Union[bytes, str]:
        """Export report data to the specific format"""
        raise NotImplementedError("Subclasses must implement export_report method")
    
    def validate_report_data(self, report_data: ReportData) -> None:
        """Validate report data before export"""
        if not report_data:
            raise ExportError("Report data is required")
        
        if not report_data.data:
            raise ExportError("Report data cannot be empty")
        
        if not report_data.summary:
            raise ExportError("Report summary is required")


class PDFExporter(BaseExporter):
    """Export reports to PDF format with professional formatting and company branding"""
    
    def __init__(self, company_data: Optional[Dict[str, Any]] = None):
        super().__init__()
        self.company_data = company_data or {}
        self.styles = getSampleStyleSheet()
        self._create_custom_styles()
    
    def _create_custom_styles(self):
        """Create custom paragraph styles for professional formatting"""
        # Title style
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Heading1'],
            fontSize=20,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2c3e50')
        ))
        
        # Company name style
        self.styles.add(ParagraphStyle(
            name='CompanyName',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            alignment=TA_LEFT,
            textColor=colors.HexColor('#34495e')
        ))
        
        # Section header style
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading3'],
            fontSize=12,
            spaceBefore=15,
            spaceAfter=8,
            textColor=colors.HexColor('#2c3e50'),
            fontName='Helvetica-Bold'
        ))
        
        # Summary style
        self.styles.add(ParagraphStyle(
            name='Summary',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=15,
            textColor=colors.HexColor('#7f8c8d')
        ))
    
    def export_report(self, report_data: ReportData, **kwargs) -> bytes:
        """Export report data to PDF format"""
        try:
            self.validate_report_data(report_data)
            
            # Create PDF buffer
            buffer = io.BytesIO()
            
            # Create document with professional margins
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            
            # Build document content
            story = []
            
            # Header with company branding
            story.extend(self._build_header(report_data))
            story.append(Spacer(1, 20))
            
            # Report summary
            story.extend(self._build_summary(report_data))
            story.append(Spacer(1, 20))
            
            # Report data table
            story.extend(self._build_data_table(report_data))
            
            # Footer
            story.extend(self._build_footer())
            
            # Build PDF
            doc.build(story)
            
            # Get PDF bytes
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            self.logger.info(f"Generated PDF report for {report_data.report_type}")
            return pdf_bytes
            
        except Exception as e:
            self.logger.error(f"Failed to generate PDF report: {str(e)}")
            raise ExportError(f"PDF export failed: {str(e)}")
    
    def _build_header(self, report_data: ReportData) -> List:
        """Build report header with company branding"""
        elements = []
        
        # Company name if available
        if self.company_data.get('name'):
            elements.append(Paragraph(self.company_data['name'], self.styles['CompanyName']))
        
        # Report title
        title = f"{report_data.report_type.value.title()} Report"
        elements.append(Paragraph(title, self.styles['ReportTitle']))
        
        # Generation date
        gen_date = report_data.metadata.generated_at.strftime('%B %d, %Y at %I:%M %p')
        elements.append(Paragraph(f"Generated on {gen_date}", self.styles['Summary']))
        
        return elements
    
    def _build_summary(self, report_data: ReportData) -> List:
        """Build report summary section"""
        elements = []
        
        elements.append(Paragraph("Report Summary", self.styles['SectionHeader']))
        
        # Summary data
        summary_data = [
            ['Total Records:', str(report_data.summary.total_records)],
            ['Date Range:', self._format_date_range(report_data.summary.date_range)],
        ]
        
        # Add total amount if available
        if report_data.summary.total_amount is not None:
            currency = report_data.summary.currency or 'USD'
            amount_str = f"{report_data.summary.total_amount:,.2f} {currency}"
            summary_data.append(['Total Amount:', amount_str])
        
        # Add key metrics
        for key, value in report_data.summary.key_metrics.items():
            if isinstance(value, (int, float)):
                if isinstance(value, float):
                    value_str = f"{value:,.2f}"
                else:
                    value_str = f"{value:,}"
            else:
                value_str = str(value)
            summary_data.append([f"{key.replace('_', ' ').title()}:", value_str])
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(summary_table)
        
        return elements
    
    def _build_data_table(self, report_data: ReportData) -> List:
        """Build main data table"""
        elements = []
        
        if not report_data.data:
            elements.append(Paragraph("No data available", self.styles['Normal']))
            return elements
        
        elements.append(Paragraph("Report Data", self.styles['SectionHeader']))
        
        # Get column headers from first row
        first_row = report_data.data[0]
        headers = list(first_row.keys())
        
        # Build table data
        table_data = [headers]  # Header row
        
        for row in report_data.data:
            table_row = []
            for header in headers:
                value = row.get(header, '')
                # Format values appropriately
                if isinstance(value, (int, float)):
                    if isinstance(value, float):
                        table_row.append(f"{value:,.2f}")
                    else:
                        table_row.append(f"{value:,}")
                elif isinstance(value, datetime):
                    table_row.append(value.strftime('%Y-%m-%d'))
                else:
                    table_row.append(str(value) if value is not None else '')
            table_data.append(table_row)
        
        # Calculate column widths dynamically
        num_cols = len(headers)
        available_width = 6.5 * inch  # Available width for table
        col_width = available_width / num_cols
        col_widths = [col_width] * num_cols
        
        # Create table
        data_table = Table(table_data, colWidths=col_widths)
        data_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows styling
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(data_table)
        
        return elements
    
    def _build_footer(self) -> List:
        """Build report footer"""
        elements = []
        
        elements.append(Spacer(1, 30))
        
        footer_text = "This report was generated automatically."
        if self.company_data.get('name'):
            footer_text = f"Generated by {self.company_data['name']} reporting system."
        
        elements.append(Paragraph(footer_text, self.styles['Summary']))
        
        return elements
    
    def _format_date_range(self, date_range: Optional[Dict[str, datetime]]) -> str:
        """Format date range for display"""
        if not date_range:
            return "All time"
        
        start_date = date_range.get('start')
        end_date = date_range.get('end')
        
        if start_date and end_date:
            return f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        elif start_date:
            return f"From {start_date.strftime('%Y-%m-%d')}"
        elif end_date:
            return f"Until {end_date.strftime('%Y-%m-%d')}"
        else:
            return "All time"


class CSVExporter(BaseExporter):
    """Export reports to CSV format for data analysis compatibility"""
    
    def export_report(self, report_data: ReportData, **kwargs) -> str:
        """Export report data to CSV format"""
        try:
            self.validate_report_data(report_data)
            
            # Create CSV buffer
            output = io.StringIO()
            
            if not report_data.data:
                return ""
            
            # Get column headers from first row
            first_row = report_data.data[0]
            headers = list(first_row.keys())
            
            # Create CSV writer
            writer = csv.DictWriter(output, fieldnames=headers)
            
            # Write header
            writer.writeheader()
            
            # Write data rows
            for row in report_data.data:
                # Convert values to strings and handle None values
                csv_row = {}
                for header in headers:
                    value = row.get(header)
                    if value is None:
                        csv_row[header] = ''
                    elif isinstance(value, datetime):
                        csv_row[header] = value.isoformat()
                    elif isinstance(value, (int, float)):
                        csv_row[header] = str(value)
                    else:
                        csv_row[header] = str(value)
                
                writer.writerow(csv_row)
            
            csv_content = output.getvalue()
            output.close()
            
            self.logger.info(f"Generated CSV report for {report_data.report_type}")
            return csv_content
            
        except Exception as e:
            self.logger.error(f"Failed to generate CSV report: {str(e)}")
            raise ExportError(f"CSV export failed: {str(e)}")


class ExcelExporter(BaseExporter):
    """Export reports to Excel format with multi-sheet support and formatting"""
    
    def __init__(self, company_data: Optional[Dict[str, Any]] = None):
        super().__init__()
        self.company_data = company_data or {}
    
    def export_report(self, report_data: ReportData, **kwargs) -> bytes:
        """Export report data to Excel format"""
        try:
            self.validate_report_data(report_data)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create main data sheet
            ws_data = wb.create_sheet(title="Report Data")
            self._build_data_sheet(ws_data, report_data)
            
            # Create summary sheet
            ws_summary = wb.create_sheet(title="Summary")
            self._build_summary_sheet(ws_summary, report_data)
            
            # Set active sheet to summary
            wb.active = ws_summary
            
            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()
            
            self.logger.info(f"Generated Excel report for {report_data.report_type}")
            return excel_bytes
            
        except Exception as e:
            self.logger.error(f"Failed to generate Excel report: {str(e)}")
            raise ExportError(f"Excel export failed: {str(e)}")
    
    def _build_summary_sheet(self, ws: Worksheet, report_data: ReportData):
        """Build summary sheet with report overview"""
        # Title
        ws['A1'] = f"{report_data.report_type.value.title()} Report Summary"
        ws['A1'].font = Font(size=16, bold=True, color='2C3E50')
        ws.merge_cells('A1:B1')
        
        # Company name if available
        row = 3
        if self.company_data.get('name'):
            ws[f'A{row}'] = "Company:"
            ws[f'B{row}'] = self.company_data['name']
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Generation date
        ws[f'A{row}'] = "Generated:"
        ws[f'B{row}'] = report_data.metadata.generated_at.strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)
        row += 2
        
        # Summary data
        ws[f'A{row}'] = "Total Records:"
        ws[f'B{row}'] = report_data.summary.total_records
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Date range
        ws[f'A{row}'] = "Date Range:"
        ws[f'B{row}'] = self._format_date_range_excel(report_data.summary.date_range)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Total amount if available
        if report_data.summary.total_amount is not None:
            ws[f'A{row}'] = "Total Amount:"
            currency = report_data.summary.currency or 'USD'
            ws[f'B{row}'] = f"{report_data.summary.total_amount:,.2f} {currency}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Key metrics
        if report_data.summary.key_metrics:
            row += 1
            ws[f'A{row}'] = "Key Metrics:"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for key, value in report_data.summary.key_metrics.items():
                ws[f'A{row}'] = f"{key.replace('_', ' ').title()}:"
                if isinstance(value, (int, float)):
                    if isinstance(value, float):
                        ws[f'B{row}'] = f"{value:,.2f}"
                    else:
                        ws[f'B{row}'] = f"{value:,}"
                else:
                    ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _build_data_sheet(self, ws: Worksheet, report_data: ReportData):
        """Build data sheet with report data"""
        if not report_data.data:
            ws['A1'] = "No data available"
            return
        
        # Get column headers from first row
        first_row = report_data.data[0]
        headers = list(first_row.keys())
        
        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data rows
        for row_idx, row_data in enumerate(report_data.data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                
                # Format values appropriately
                if value is None:
                    cell_value = ''
                elif isinstance(value, datetime):
                    cell_value = value
                elif isinstance(value, (int, float)):
                    cell_value = value
                else:
                    cell_value = str(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Apply alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(report_data.data) + 1, 
                               min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _format_date_range_excel(self, date_range: Optional[Dict[str, datetime]]) -> str:
        """Format date range for Excel display"""
        if not date_range:
            return "All time"
        
        start_date = date_range.get('start')
        end_date = date_range.get('end')
        
        if start_date and end_date:
            return f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        elif start_date:
            return f"From {start_date.strftime('%Y-%m-%d')}"
        elif end_date:
            return f"Until {end_date.strftime('%Y-%m-%d')}"
        else:
            return "All time"


class ReportExportService:
    """Main service for handling report exports in multiple formats"""
    
    def __init__(self, company_data: Optional[Dict[str, Any]] = None):
        self.company_data = company_data or {}
        self.logger = logging.getLogger(__name__)
        
        # Initialize exporters
        self.exporters = {
            ExportFormat.PDF: PDFExporter(company_data),
            ExportFormat.CSV: CSVExporter(),
            ExportFormat.EXCEL: ExcelExporter(company_data),
        }
    
    def export_report(
        self, 
        report_data: ReportData, 
        export_format: ExportFormat,
        **kwargs
    ) -> Union[bytes, str]:
        """Export report in the specified format"""
        try:
            # Validate export format
            if export_format not in self.exporters:
                raise ExportError(f"Unsupported export format: {export_format}")
            
            # Get appropriate exporter
            exporter = self.exporters[export_format]
            
            # Export report
            result = exporter.export_report(report_data, **kwargs)
            
            self.logger.info(f"Successfully exported {report_data.report_type} report as {export_format}")
            return result
            
        except Exception as e:
            self.logger.error(f"Export failed for format {export_format}: {str(e)}")
            raise ExportError(f"Export failed: {str(e)}")
    
    def get_supported_formats(self) -> List[ExportFormat]:
        """Get list of supported export formats"""
        return list(self.exporters.keys())
    
    def validate_export_format(self, export_format: str) -> ExportFormat:
        """Validate and convert export format string to enum"""
        try:
            return ExportFormat(export_format.lower())
        except ValueError:
            supported = [fmt.value for fmt in self.get_supported_formats()]
            raise ExportError(f"Invalid export format '{export_format}'. Supported formats: {supported}")