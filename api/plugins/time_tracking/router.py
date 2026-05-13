"""
Time Tracking Plugin — FastAPI Routers

Two APIRouter instances:
  - projects_router  (mounted at /api/v1/projects)
  - time_entries_router (mounted at /api/v1/time-entries)

All routes require authentication via `get_current_user`.
Multi-tenant isolation is via `get_db` which returns the tenant-specific DB session.

Excel export uses openpyxl, matching the pattern in core/services/report_exporter.py.
"""

from __future__ import annotations

import io
import csv
import json
import logging
from datetime import datetime, time, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from core.models.database import get_db
from core.models.models import MasterUser
from core.models.models_per_tenant import Client, Invoice, InvoiceItem, Expense
from core.routers.auth import get_current_user
from core.utils.audit import log_audit_event

from .models import Project, ProjectTask, TimeEntry
from .schemas import (
    ProjectCreate, ProjectUpdate, ProjectResponse,
    ProjectTaskCreate, ProjectTaskUpdate, ProjectTaskResponse,
    TimeEntryCreate, TimeEntryUpdate, TimeEntryResponse,
    TimerStartRequest, TimerStopRequest, TimerActiveResponse,
    ProjectSummaryResponse, UnbilledItemsResponse,
    UnbilledTimeEntry, UnbilledExpense,
    ProjectInvoiceRequest, ProjectInvoiceResponse,
    TimeExportFilters, TimeExportRow,
    TimeImportError, TimeImportResponse,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _enrich_project(project: Project, db: Session) -> dict:
    """Add client_name and aggregated stats to a project dict."""
    client = db.query(Client).filter(Client.id == project.client_id).first()
    hours_agg = (
        db.query(func.sum(TimeEntry.duration_minutes))
        .filter(TimeEntry.project_id == project.id, TimeEntry.status != "in_progress")
        .scalar()
    ) or 0
    amount_agg = (
        db.query(func.sum(TimeEntry.amount))
        .filter(TimeEntry.project_id == project.id, TimeEntry.invoiced.is_(False))
        .scalar()
    ) or 0.0

    data = {
        "id": project.id,
        "client_id": project.client_id,
        "name": project.name,
        "description": project.description,
        "billing_method": project.billing_method,
        "hourly_rate": project.hourly_rate,
        "fixed_amount": project.fixed_amount,
        "budget_hours": project.budget_hours,
        "budget_amount": project.budget_amount,
        "status": project.status,
        "currency": project.currency,
        "created_by": project.created_by,
        "created_at": project.created_at,
        "updated_at": project.updated_at,
        "client_name": client.name if client else None,
        "total_hours_logged": round(hours_agg / 60.0, 2) if hours_agg else 0.0,
        "total_amount_logged": round(float(amount_agg), 2),
    }
    return data


def _enrich_time_entry(entry: TimeEntry, db: Session) -> dict:
    """Add project_name, task_name, client_name to a time entry dict."""
    project = db.query(Project).filter(Project.id == entry.project_id).first()
    task = db.query(ProjectTask).filter(ProjectTask.id == entry.task_id).first() if entry.task_id else None
    client = db.query(Client).filter(Client.id == entry.client_id).first() if entry.client_id else None
    return {
        "id": entry.id,
        "project_id": entry.project_id,
        "task_id": entry.task_id,
        "user_id": entry.user_id,
        "client_id": entry.client_id,
        "description": entry.description,
        "notes": entry.notes,
        "started_at": entry.started_at,
        "ended_at": entry.ended_at,
        "duration_minutes": entry.duration_minutes,
        "hourly_rate": entry.hourly_rate,
        "billable": entry.billable,
        "amount": entry.amount,
        "status": entry.status,
        "invoiced": entry.invoiced,
        "invoice_id": entry.invoice_id,
        "invoice_number": entry.invoice_number,
        "created_at": entry.created_at,
        "updated_at": entry.updated_at,
        "hours": entry.hours,
        "project_name": project.name if project else None,
        "task_name": task.name if task else None,
        "client_name": client.name if client else None,
    }


TIME_IMPORT_FIELD_ALIASES = {
    "client_id": ["client_id", "client id"],
    "client_name": ["client", "client_name", "client name", "customer", "customer_name", "customer name"],
    "project_id": ["project_id", "project id"],
    "project_name": ["project", "project_name", "project name", "job", "job_name", "job name"],
    "task_name": ["task", "task_name", "task name", "activity", "service"],
    "description": ["description", "desc", "work", "work performed", "memo"],
    "notes": ["notes", "note"],
    "date": ["date", "work_date", "work date", "entry_date", "entry date"],
    "started_at": ["started_at", "started at", "start_datetime", "start datetime", "start date"],
    "ended_at": ["ended_at", "ended at", "end_datetime", "end datetime", "end date"],
    "start_time": ["start", "start_time", "start time", "from"],
    "end_time": ["end", "end_time", "end time", "to"],
    "duration_minutes": ["duration_minutes", "duration minutes", "minutes", "mins"],
    "hours": ["hours", "hrs", "duration_hours", "duration hours", "time", "duration"],
    "hourly_rate": ["hourly_rate", "hourly rate", "rate", "bill rate", "billing rate"],
    "billable": ["billable", "is_billable", "is billable"],
    "currency": ["currency"],
    "billing_method": ["billing_method", "billing method"],
}


def _parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip().replace("$", "").replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_int(value: Any) -> Optional[int]:
    parsed = _parse_float(value)
    return int(parsed) if parsed is not None else None


def _parse_bool(value: Any, default: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "y", "billable"}


def _parse_datetime(value: Any) -> Optional[datetime]:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    for candidate in (text, text.replace("Z", "+00:00")):
        try:
            parsed = datetime.fromisoformat(candidate)
            return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %I:%M %p", "%m/%d/%Y %H:%M", "%m/%d/%Y %I:%M %p", "%m/%d/%y %I:%M %p"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return None


def _parse_date(value: Any) -> Optional[datetime]:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.combine(datetime.strptime(text, fmt).date(), time.min, tzinfo=timezone.utc)
        except ValueError:
            pass
    return _parse_datetime(text)


def _parse_time(value: Any) -> Optional[time]:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I %p"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    return None


def _canonicalize_csv_rows(csv_content: str, field_mapping: Optional[Dict[str, str]] = None) -> List[Dict[str, Any]]:
    reader = csv.DictReader(io.StringIO(csv_content))
    if not reader.fieldnames:
        return []

    header_map = {header: header.strip().lower().replace("_", " ") for header in reader.fieldnames if header}
    rows: List[Dict[str, Any]] = []
    for raw in reader:
        normalized: Dict[str, Any] = {}
        for canonical, aliases in TIME_IMPORT_FIELD_ALIASES.items():
            for source, lowered in header_map.items():
                if lowered in aliases:
                    normalized[canonical] = raw.get(source)
                    break
        for canonical, source in (field_mapping or {}).items():
            if canonical in TIME_IMPORT_FIELD_ALIASES and source in raw:
                normalized[canonical] = raw.get(source)
        rows.append(normalized)
    return rows


def _extract_json_object(text: str) -> Dict[str, Any]:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.split("```", 2)[1]
        if cleaned.startswith("json"):
            cleaned = cleaned[4:]
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            return json.loads(cleaned[start:end + 1])
        raise


def _normalize_rows_with_ai(csv_content: str, db: Session) -> Optional[List[Dict[str, Any]]]:
    try:
        from commercial.ai.services.ai_config_service import AIConfigService

        ai_config = AIConfigService.get_ai_config(db, component="ocr", require_ocr=False)
        if not ai_config:
            return None

        provider = (ai_config.get("provider_name") or "").lower()
        model_name = ai_config.get("model_name")
        reader = csv.DictReader(io.StringIO(csv_content))
        headers = reader.fieldnames or []
        sample_rows = []
        for _, row in zip(range(5), reader):
            sample_rows.append(row)

        prompt = f"""
Map this CSV's column headers to time tracking import fields. Return JSON only.

Return this exact shape:
{{"mapping":{{"client_id":null,"client_name":null,"project_id":null,"project_name":null,"task_name":null,"description":null,"notes":null,"date":null,"started_at":null,"ended_at":null,"start_time":null,"end_time":null,"duration_minutes":null,"hours":null,"hourly_rate":null,"billable":null,"currency":null,"billing_method":null}}}}

Rules:
- Each mapping value must be one exact source header from the CSV, or null.
- Prefer project_id/client_id over names only when the source column contains numeric IDs.
- Use date plus start_time/end_time when the CSV has separate date and time columns.
- Use started_at/ended_at when the CSV has full datetime columns.
- Do not invent headers that are not in the CSV.

CSV headers:
{json.dumps(headers)}

Sample rows:
{json.dumps(sample_rows, default=str)}
"""
        if provider == "openai":
            from langchain_openai import ChatOpenAI

            llm = ChatOpenAI(
                api_key=ai_config.get("api_key"),
                model=model_name,
                temperature=0.0,
                max_tokens=6000,
                model_kwargs={"response_format": {"type": "json_object"}},
            )
        elif provider == "anthropic":
            from langchain_anthropic import ChatAnthropic

            llm = ChatAnthropic(
                api_key=ai_config.get("api_key"),
                model=model_name,
                temperature=0.0,
                max_tokens=6000,
            )
        elif provider == "ollama":
            from langchain_ollama import OllamaLLM

            llm = OllamaLLM(
                base_url=ai_config.get("provider_url", "http://localhost:11434"),
                model=model_name,
                temperature=0.0,
                num_predict=6000,
            )
        else:
            return None

        response = llm.invoke(prompt)
        response_text = response.content if hasattr(response, "content") else str(response)
        payload = _extract_json_object(response_text)
        raw_mapping = payload.get("mapping")
        if not isinstance(raw_mapping, dict):
            return None

        valid_headers = set(headers)
        field_mapping = {
            str(canonical): str(source)
            for canonical, source in raw_mapping.items()
            if source in valid_headers and canonical in TIME_IMPORT_FIELD_ALIASES
        }
        if not field_mapping:
            return None
        logger.info("AI time import header mapping: %s", field_mapping)
        return _canonicalize_csv_rows(csv_content, field_mapping)
    except Exception as exc:
        logger.warning("AI time import normalization failed: %s", exc)
        return None


def _find_client_by_name(db: Session, name: str) -> Optional[Client]:
    target = name.strip().lower()
    for client in db.query(Client).all():
        if (client.name or "").strip().lower() == target:
            return client
    return None


def _find_project(db: Session, project_name: str, client_id: int) -> Optional[Project]:
    target = project_name.strip().lower()
    for project in db.query(Project).filter(Project.client_id == client_id).all():
        if project.name.strip().lower() == target:
            return project
    return None


def _find_project_by_name(db: Session, project_name: str) -> Optional[Project]:
    target = project_name.strip().lower()
    for project in db.query(Project).all():
        if project.name.strip().lower() == target:
            return project
    return None


def _get_or_create_import_client(db: Session, current_user: MasterUser, now: datetime, currency: str) -> tuple[Client, bool]:
    client = _find_client_by_name(db, "Imported Time")
    if client:
        return client, False

    client = Client(
        name="Imported Time",
        preferred_currency=currency,
        source="time_import",
        owner_user_id=current_user.id,
        created_at=now,
        updated_at=now,
    )
    db.add(client)
    db.flush()
    return client, True


def _find_task(db: Session, project_id: int, task_name: str) -> Optional[ProjectTask]:
    target = task_name.strip().lower()
    for task in db.query(ProjectTask).filter(ProjectTask.project_id == project_id).all():
        if task.name.strip().lower() == target:
            return task
    return None


def _resolve_hourly_rate(row_rate: Optional[float], task: Optional[ProjectTask], project: Project) -> float:
    if row_rate is not None:
        return row_rate
    if task and task.hourly_rate is not None:
        return task.hourly_rate
    if project.hourly_rate is not None:
        return project.hourly_rate
    return 0.0


# ---------------------------------------------------------------------------
# Projects Router
# ---------------------------------------------------------------------------

projects_router = APIRouter()


@projects_router.post("", response_model=ProjectResponse, status_code=201)
def create_project(
    payload: ProjectCreate,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """Create a new project linked to a client."""
    client = db.query(Client).filter(Client.id == payload.client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    project = Project(
        **payload.model_dump(),
        created_by=current_user.id,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    )
    db.add(project)
    db.commit()
    db.refresh(project)

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="CREATE",
        resource_type="project",
        resource_id=str(project.id),
        resource_name=project.name,
        details=payload.model_dump(),
        status="success"
    )

    return _enrich_project(project, db)


@projects_router.get("", response_model=List[ProjectResponse])
def list_projects(
    status: Optional[str] = Query(None),
    client_id: Optional[int] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """List all projects with optional filters."""
    q = db.query(Project)
    if status:
        q = q.filter(Project.status == status)
    if client_id:
        q = q.filter(Project.client_id == client_id)
    projects = q.order_by(Project.created_at.desc()).offset(skip).limit(limit).all()
    return [_enrich_project(p, db) for p in projects]


@projects_router.get("/{project_id}", response_model=ProjectResponse)
def get_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return _enrich_project(project, db)


@projects_router.patch("/{project_id}", response_model=ProjectResponse)
def update_project(
    project_id: int,
    payload: ProjectUpdate,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    for field, value in payload.model_dump(exclude_unset=True).items():
        if field == "client_id" and value is not None:
            client = db.query(Client).filter(Client.id == value).first()
            if not client:
                raise HTTPException(status_code=404, detail="Client not found")
        setattr(project, field, value)
    project.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(project)

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="UPDATE",
        resource_type="project",
        resource_id=str(project.id),
        resource_name=project.name,
        details=payload.model_dump(exclude_unset=True),
        status="success"
    )

    return _enrich_project(project, db)


@projects_router.delete("/{project_id}", status_code=204)
def delete_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """Soft-delete: sets status to 'archived'."""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    project.status = "archived"
    project.updated_at = datetime.now(timezone.utc)
    db.commit()

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="DELETE",
        resource_type="project",
        resource_id=str(project.id),
        resource_name=project.name,
        details={"status": "archived"},
        status="success"
    )


@projects_router.get("/{project_id}/summary", response_model=ProjectSummaryResponse)
def get_project_summary(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """Return KPI summary for a project."""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    client = db.query(Client).filter(Client.id == project.client_id).first()

    # Aggregate time entries (logged/approved only, not in_progress)
    total_minutes = (
        db.query(func.sum(TimeEntry.duration_minutes))
        .filter(TimeEntry.project_id == project_id, TimeEntry.status != "in_progress")
        .scalar()
    ) or 0
    total_amount = (
        db.query(func.sum(TimeEntry.amount))
        .filter(TimeEntry.project_id == project_id, TimeEntry.status != "in_progress")
        .scalar()
    ) or 0.0
    total_expenses = (
        db.query(func.sum(Expense.amount))
        .filter(getattr(Expense, "project_id", None) == project_id)
        .scalar()
    ) or 0.0

    # Unbilled
    unbilled_minutes = (
        db.query(func.sum(TimeEntry.duration_minutes))
        .filter(
            TimeEntry.project_id == project_id,
            TimeEntry.invoiced.is_(False),
            TimeEntry.status != "in_progress",
        )
        .scalar()
    ) or 0
    unbilled_amount = (
        db.query(func.sum(TimeEntry.amount))
        .filter(
            TimeEntry.project_id == project_id,
            TimeEntry.invoiced.is_(False),
            TimeEntry.status != "in_progress",
        )
        .scalar()
    ) or 0.0

    total_hours = round(total_minutes / 60.0, 2)
    unbilled_hours = round(unbilled_minutes / 60.0, 2)

    hours_pct = None
    if project.budget_hours and project.budget_hours > 0:
        hours_pct = round((total_hours / project.budget_hours) * 100, 1)

    budget_pct = None
    if project.budget_amount and project.budget_amount > 0:
        budget_pct = round((float(total_amount) / project.budget_amount) * 100, 1)

    return ProjectSummaryResponse(
        project_id=project.id,
        project_name=project.name,
        client_id=project.client_id,
        client_name=client.name if client else None,
        status=project.status,
        billing_method=project.billing_method,
        budget_hours=project.budget_hours,
        budget_amount=project.budget_amount,
        total_hours_logged=total_hours,
        total_amount_logged=round(float(total_amount), 2),
        total_expenses=round(float(total_expenses), 2),
        unbilled_hours=unbilled_hours,
        unbilled_amount=round(float(unbilled_amount), 2),
        hours_used_pct=hours_pct,
        budget_used_pct=budget_pct,
    )


@projects_router.get("/{project_id}/unbilled", response_model=UnbilledItemsResponse)
def get_unbilled_items(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """Return all unbilled time entries and expenses for a project."""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    entries = (
        db.query(TimeEntry)
        .filter(
            TimeEntry.project_id == project_id,
            TimeEntry.invoiced.is_(False),
            TimeEntry.status != "in_progress",
        )
        .all()
    )

    # Expenses tagged to this project — handles case where project_id column may not exist
    expenses = []
    try:
        expenses = (
            db.query(Expense)
            .filter(
                getattr(Expense, "project_id") == project_id,  # noqa: B009
                getattr(Expense, "invoiced", False) == False,  # noqa: E712,B009
            )
            .all()
        )
    except Exception:
        pass

    time_items = []
    for e in entries:
        task = db.query(ProjectTask).filter(ProjectTask.id == e.task_id).first() if e.task_id else None
        time_items.append(
            UnbilledTimeEntry(
                id=e.id,
                task_name=task.name if task else None,
                description=e.description,
                started_at=e.started_at,
                hours=e.hours,
                hourly_rate=e.hourly_rate,
                amount=e.amount or 0.0,
                billable=e.billable,
            )
        )

    expense_items = [
        UnbilledExpense(
            id=exp.id,
            category=getattr(exp, "category", None),
            vendor=getattr(exp, "vendor", None),
            expense_date=str(getattr(exp, "expense_date", "")),
            amount=float(getattr(exp, "amount", 0) or 0),
            currency=getattr(exp, "currency", None),
        )
        for exp in expenses
    ]

    total_time = sum(i.amount for i in time_items)
    total_expense = sum(i.amount for i in expense_items)

    return UnbilledItemsResponse(
        project_id=project_id,
        time_entries=time_items,
        expenses=expense_items,
        total_time_amount=round(total_time, 2),
        total_expense_amount=round(total_expense, 2),
        grand_total=round(total_time + total_expense, 2),
    )


@projects_router.post("/{project_id}/invoice", response_model=ProjectInvoiceResponse, status_code=201)
def create_invoice_from_project(
    project_id: int,
    payload: ProjectInvoiceRequest,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """Generate an invoice from selected unbilled time entries and expenses."""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if not payload.time_entry_ids and not payload.expense_ids:
        raise HTTPException(status_code=400, detail="No items selected for invoicing")

    # Fetch selected time entries
    time_entries = (
        db.query(TimeEntry)
        .filter(
            TimeEntry.id.in_(payload.time_entry_ids),
            TimeEntry.project_id == project_id,
            TimeEntry.invoiced.is_(False),
        )
        .all()
    )

    expenses = []
    if payload.expense_ids:
        try:
            expenses = (
                db.query(Expense)
                .filter(
                    Expense.id.in_(payload.expense_ids),
                    getattr(Expense, "project_id") == project_id,  # noqa: B009
                )
                .all()
            )
        except Exception:
            pass

    # Build invoice line items
    line_items = []
    total = 0.0

    for entry in time_entries:
        amount = entry.amount or 0.0
        line_items.append({
            "description": f"{entry.description or 'Time'} ({entry.hours:.2f}h @ {entry.hourly_rate}/hr)",
            "quantity": entry.hours,
            "price": entry.hourly_rate,
            "amount": amount,
        })
        total += amount

    for exp in expenses:
        amt = float(getattr(exp, "amount", 0) or 0)
        line_items.append({
            "description": f"Expense: {getattr(exp, 'vendor', '') or ''} - {getattr(exp, 'category', '')}",
            "quantity": 1,
            "price": amt,
            "amount": amt,
        })
        total += amt

    if not line_items:
        raise HTTPException(status_code=400, detail="No valid items found to invoice")

    # Generate invoice number
    existing_count = db.query(Invoice).count()
    invoice_number = f"INV-{existing_count + 1:05d}"

    now = datetime.now(timezone.utc)
    due_date_str = payload.due_date or now.strftime("%Y-%m-%d")
    due_date = datetime.strptime(due_date_str, "%Y-%m-%d") if isinstance(due_date_str, str) else due_date_str

    invoice = Invoice(
        number=invoice_number,
        client_id=project.client_id,
        due_date=due_date,
        amount=round(total, 2),
        subtotal=round(total, 2),
        currency=project.currency,
        status="pending",
        notes=payload.notes or f"Generated from project: {project.name}",
        created_at=now,
        updated_at=now,
        created_by_user_id=current_user.id
    )
    db.add(invoice)
    db.flush()  # get invoice.id

    # Add line items
    for item in line_items:
        db.add(InvoiceItem(
            invoice_id=invoice.id,
            description=item["description"],
            quantity=item["quantity"],
            price=item["price"],
            amount=item["amount"],
        ))

    # Mark time entries as invoiced
    for entry in time_entries:
        entry.invoiced = True
        entry.invoice_id = invoice.id
        entry.invoice_number = invoice_number
        entry.status = "invoiced"
        entry.updated_at = now

    # Mark expenses as invoiced (if column exists)
    for exp in expenses:
        try:
            exp.invoiced = True
            exp.invoice_id = invoice.id
        except Exception:
            pass

    db.commit()
    db.refresh(invoice)

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="CREATE",
        resource_type="invoice",
        resource_id=str(invoice.id),
        resource_name=f"Invoice {invoice.number}",
        details={
            "project_id": project.id,
            "amount": invoice.amount,
            "currency": invoice.currency,
            "due_date": due_date_str,
            "line_items_count": len(line_items)
        },
        status="success"
    )

    return ProjectInvoiceResponse(
        invoice_id=invoice.id,
        invoice_number=invoice.number,
        amount=invoice.amount,
        currency=invoice.currency,
    )


# ---------------------------------------------------------------------------
# Task sub-routes (nested under projects)
# ---------------------------------------------------------------------------

@projects_router.post("/{project_id}/tasks", response_model=ProjectTaskResponse, status_code=201)
def create_task(
    project_id: int,
    payload: ProjectTaskCreate,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    task = ProjectTask(
        project_id=project_id,
        name=payload.name,
        description=payload.description,
        estimated_hours=payload.estimated_hours,
        hourly_rate=payload.hourly_rate,
        status=payload.status,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    )
    db.add(task)
    db.commit()
    db.refresh(task)

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="CREATE",
        resource_type="project_task",
        resource_id=str(task.id),
        resource_name=task.name,
        details=payload.model_dump(),
        status="success"
    )

    # Inject the computed field directly on the ORM object.
    # Spreading task.__dict__ after db.refresh() only contains _sa_instance_state
    # (all column attrs are expired), causing a Pydantic ResponseValidationError.
    # With from_attributes=True, Pydantic reads real column values on demand.
    task.actual_hours = 0.0
    return task


@projects_router.get("/{project_id}/tasks", response_model=List[ProjectTaskResponse])
def list_tasks(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    tasks = db.query(ProjectTask).filter(ProjectTask.project_id == project_id).all()
    results = []
    for task in tasks:
        actual_minutes = (
            db.query(func.sum(TimeEntry.duration_minutes))
            .filter(TimeEntry.task_id == task.id, TimeEntry.status != "in_progress")
            .scalar()
        ) or 0
        results.append({**task.__dict__, "actual_hours": round(actual_minutes / 60.0, 2)})
    return results


@projects_router.patch("/{project_id}/tasks/{task_id}", response_model=ProjectTaskResponse)
def update_task(
    project_id: int,
    task_id: int,
    payload: ProjectTaskUpdate,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    task = db.query(ProjectTask).filter(
        ProjectTask.id == task_id,
        ProjectTask.project_id == project_id,
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(task, field, value)
    task.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(task)

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="UPDATE",
        resource_type="project_task",
        resource_id=str(task.id),
        resource_name=task.name,
        details=payload.model_dump(exclude_unset=True),
        status="success"
    )

    actual_minutes = (
        db.query(func.sum(TimeEntry.duration_minutes))
        .filter(TimeEntry.task_id == task_id, TimeEntry.status != "in_progress")
        .scalar()
    ) or 0
    task.actual_hours = round(actual_minutes / 60.0, 2)
    return task


@projects_router.delete("/{project_id}/tasks/{task_id}", status_code=204)
def delete_task(
    project_id: int,
    task_id: int,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    task = db.query(ProjectTask).filter(
        ProjectTask.id == task_id,
        ProjectTask.project_id == project_id,
    ).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    db.delete(task)
    db.commit()

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="DELETE",
        resource_type="project_task",
        resource_id=str(task_id),
        resource_name=task.name,
        status="success"
    )


# ---------------------------------------------------------------------------
# Time Entries Router
# ---------------------------------------------------------------------------

time_entries_router = APIRouter()


# ---- IMPORTANT: static paths first, THEN /{entry_id} ----

@time_entries_router.post("/timer/start", response_model=TimeEntryResponse, status_code=201)
def timer_start(
    payload: TimerStartRequest,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """Start the live timer. Only one active timer per user is allowed."""
    # Check for an already-running timer
    existing = (
        db.query(TimeEntry)
        .filter(TimeEntry.user_id == current_user.id, TimeEntry.status == "in_progress")
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail="A timer is already running. Stop it first.")

    project = db.query(Project).filter(Project.id == payload.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    task = db.query(ProjectTask).filter(ProjectTask.id == payload.task_id).first() if payload.task_id else None
    now = datetime.now(timezone.utc)
    started = payload.started_at or now
    hourly_rate = _resolve_hourly_rate(payload.hourly_rate if payload.hourly_rate > 0 else None, task, project)

    entry = TimeEntry(
        project_id=payload.project_id,
        task_id=payload.task_id,
        user_id=current_user.id,
        client_id=project.client_id,
        description=payload.description,
        hourly_rate=hourly_rate,
        billable=payload.billable,
        started_at=started,
        ended_at=None,
        status="in_progress",
        invoiced=False,
        created_at=now,
        updated_at=now,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="START",
        resource_type="timer",
        resource_id=str(entry.id),
        resource_name=f"Timer for Project {payload.project_id}",
        details=payload.model_dump(),
        status="success"
    )

    return _enrich_time_entry(entry, db)


@time_entries_router.post("/timer/stop", response_model=TimeEntryResponse)
def timer_stop(
    payload: TimerStopRequest,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """Stop the active timer, compute duration and amount."""
    entry = (
        db.query(TimeEntry)
        .filter(TimeEntry.user_id == current_user.id, TimeEntry.status == "in_progress")
        .first()
    )
    if not entry:
        raise HTTPException(status_code=404, detail="No active timer found")

    now = datetime.now(timezone.utc)
    ended = payload.ended_at or now

    # Compute duration
    delta = ended - entry.started_at
    entry.ended_at = ended
    entry.duration_minutes = max(1, int(delta.total_seconds() / 60))
    entry.notes = payload.notes
    entry.status = "logged"
    entry.updated_at = now
    entry.compute_amount()

    db.commit()
    db.refresh(entry)

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="STOP",
        resource_type="timer",
        resource_id=str(entry.id),
        resource_name=f"Timer for Project {entry.project_id}",
        details={
            "duration_minutes": entry.duration_minutes,
            "amount": entry.amount,
            "notes": entry.notes
        },
        status="success"
    )

    return _enrich_time_entry(entry, db)


@time_entries_router.get("/timer/active", response_model=TimerActiveResponse)
def timer_active(
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """Return the currently active timer for the current user (if any)."""
    entry = (
        db.query(TimeEntry)
        .filter(TimeEntry.user_id == current_user.id, TimeEntry.status == "in_progress")
        .order_by(TimeEntry.started_at.desc())
        .first()
    )
    if not entry:
        return TimerActiveResponse(active=False, entry=None, elapsed_seconds=None)

    now = datetime.now(timezone.utc)
    started = entry.started_at
    if started.tzinfo is None:
        started = started.replace(tzinfo=timezone.utc)
    elapsed = int((now - started).total_seconds())

    enriched = _enrich_time_entry(entry, db)
    return TimerActiveResponse(
        active=True,
        entry=TimeEntryResponse(**enriched),
        elapsed_seconds=elapsed,
    )


@time_entries_router.get("/export/monthly")
def export_monthly(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    project_id: Optional[int] = Query(None),
    client_id: Optional[int] = Query(None),
    user_id: Optional[int] = Query(None),
    billable_only: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """
    Export a monthly time report as a .xlsx file.
    Sheet 1: Time Log (one row per entry)
    Sheet 2: Summary (totals by project)
    """
    from calendar import monthrange
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    _, last_day = monthrange(year, month)
    start_dt = datetime(year, month, 1, tzinfo=timezone.utc)
    end_dt = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)

    q = (
        db.query(TimeEntry)
        .filter(
            TimeEntry.started_at >= start_dt,
            TimeEntry.started_at <= end_dt,
            TimeEntry.status != "in_progress",
        )
    )
    if project_id:
        q = q.filter(TimeEntry.project_id == project_id)
    if client_id:
        q = q.filter(TimeEntry.client_id == client_id)
    if user_id:
        q = q.filter(TimeEntry.user_id == user_id)
    if billable_only:
        q = q.filter(TimeEntry.billable.is_(True))

    entries = q.order_by(TimeEntry.started_at.asc()).all()

    # Build row data
    rows: list[TimeExportRow] = []
    project_totals: dict[str, dict] = {}

    for entry in entries:
        project = db.query(Project).filter(Project.id == entry.project_id).first()
        task = db.query(ProjectTask).filter(ProjectTask.id == entry.task_id).first() if entry.task_id else None
        client = db.query(Client).filter(Client.id == entry.client_id).first() if entry.client_id else None

        proj_name = project.name if project else f"Project {entry.project_id}"
        client_name = client.name if client else f"Client {entry.client_id}"

        row = TimeExportRow(
            date=entry.started_at.strftime("%Y-%m-%d"),
            client_id=entry.client_id,
            client_name=client_name,
            project_name=proj_name,
            task_name=task.name if task else None,
            description=entry.description,
            notes=entry.notes,
            hours=entry.hours,
            hourly_rate=entry.hourly_rate,
            amount=entry.amount or 0.0,
            billable=entry.billable,
            status=entry.status,
            invoiced=entry.invoiced,
            invoice_number=entry.invoice_number,
        )
        rows.append(row)

        # Accumulate summary by project
        if proj_name not in project_totals:
            project_totals[proj_name] = {"hours": 0.0, "amount": 0.0, "client": client_name}
        project_totals[proj_name]["hours"] += entry.hours
        project_totals[proj_name]["amount"] += entry.amount or 0.0

    # Build Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # --- Sheet 1: Time Log ---
    ws = wb.create_sheet(title="Time Log")
    headers = [
        "Date", "Client ID", "Client Name", "Project", "Task",
        "Description", "Notes", "Hours", "Hourly Rate", "Amount",
        "Billable", "Status", "Invoiced", "Invoice #"
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.date)
        ws.cell(row=row_idx, column=2, value=row.client_id)
        ws.cell(row=row_idx, column=3, value=row.client_name)
        ws.cell(row=row_idx, column=4, value=row.project_name)
        ws.cell(row=row_idx, column=5, value=row.task_name)
        ws.cell(row=row_idx, column=6, value=row.description)
        ws.cell(row=row_idx, column=7, value=row.notes)
        ws.cell(row=row_idx, column=8, value=round(row.hours, 2))
        ws.cell(row=row_idx, column=9, value=row.hourly_rate)
        ws.cell(row=row_idx, column=10, value=round(row.amount, 2))
        ws.cell(row=row_idx, column=11, value="Yes" if row.billable else "No")
        ws.cell(row=row_idx, column=12, value=row.status)
        ws.cell(row=row_idx, column=13, value="Yes" if row.invoiced else "No")
        ws.cell(row=row_idx, column=14, value=row.invoice_number)

        # Alternating row colors
        if row_idx % 2 == 0:
            row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # Totals row
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=7, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=round(sum(r.hours for r in rows), 2)).font = Font(bold=True)
    ws.cell(row=total_row, column=10, value=round(sum(r.amount for r in rows), 2)).font = Font(bold=True)

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet(title="Summary")
    ws2["A1"] = f"Time Report — {year}-{month:02d}"
    ws2["A1"].font = Font(size=14, bold=True)

    ws2["A3"] = "Project"
    ws2["B3"] = "Client"
    ws2["C3"] = "Total Hours"
    ws2["D3"] = "Total Amount"
    for col in ["A3", "B3", "C3", "D3"]:
        ws2[col].font = Font(bold=True)
        ws2[col].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        ws2[col].font = Font(bold=True, color="FFFFFF")

    for i, (proj_name, data) in enumerate(project_totals.items(), 4):
        ws2.cell(row=i, column=1, value=proj_name)
        ws2.cell(row=i, column=2, value=data["client"])
        ws2.cell(row=i, column=3, value=round(data["hours"], 2))
        ws2.cell(row=i, column=4, value=round(data["amount"], 2))

    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws2.column_dimensions[col_letter].width = min(max_len + 2, 40)

    # Serialize to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_bytes = buffer.getvalue()
    buffer.close()

    filename = f"time_report_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@time_entries_router.post("/import/csv", response_model=TimeImportResponse, status_code=201)
async def import_time_entries_csv(
    file: UploadFile = File(...),
    use_ai: bool = Form(False),
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """
    Import projects, optional tasks, and logged time entries from a CSV file.

    Recognized columns include client/client_name/client_id, project/project_name/project_id,
    task/task_name, date, start/end or started_at/ended_at, hours/duration_minutes,
    hourly_rate/rate, billable, description, notes, currency, and billing_method.
    If use_ai is true and an AI provider is configured, the model first normalizes
    unusual CSV headers into that canonical shape.
    """
    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Please upload a CSV file")

    content = await file.read()
    if len(content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="CSV import is limited to 2MB")

    try:
        csv_content = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="CSV file must be UTF-8 encoded")

    ai_rows = _normalize_rows_with_ai(csv_content, db) if use_ai else None
    rows = ai_rows or _canonicalize_csv_rows(csv_content)
    if not rows:
        raise HTTPException(status_code=400, detail="CSV file has no importable rows")

    result = TimeImportResponse(ai_used=bool(ai_rows))
    now = datetime.now(timezone.utc)
    created_project_ids: set[int] = set()
    reused_project_ids: set[int] = set()
    created_task_ids: set[int] = set()

    for index, row in enumerate(rows, start=2):
        try:
            started_at = _parse_datetime(row.get("started_at"))
            ended_at = _parse_datetime(row.get("ended_at"))
            if not started_at:
                date_dt = _parse_date(row.get("date"))
                start_time = _parse_time(row.get("start_time")) or time(0, 0)
                if date_dt:
                    started_at = datetime.combine(date_dt.date(), start_time, tzinfo=timezone.utc)
            if started_at and not ended_at:
                end_time = _parse_time(row.get("end_time"))
                if end_time:
                    ended_at = datetime.combine(started_at.date(), end_time, tzinfo=timezone.utc)
                    if ended_at < started_at:
                        ended_at += timedelta(days=1)

            duration_minutes = _parse_int(row.get("duration_minutes"))
            hours = _parse_float(row.get("hours"))
            if duration_minutes is None and hours is not None:
                duration_minutes = max(1, int(round(hours * 60)))
            if duration_minutes is None and started_at and ended_at:
                duration_minutes = max(1, int((ended_at - started_at).total_seconds() / 60))
            if not ended_at and started_at and duration_minutes:
                ended_at = started_at + timedelta(minutes=duration_minutes)

            if not started_at:
                raise ValueError("A date/start time or started_at value is required")
            if duration_minutes is None:
                raise ValueError("Duration is required as hours, duration_minutes, or end time")

            project_id = _parse_int(row.get("project_id"))
            project_name = str(row.get("project_name") or "").strip()
            if project_id:
                project = db.query(Project).filter(Project.id == project_id).first()
                if not project:
                    raise ValueError(f"Project #{project_id} was not found")
                client_id = project.client_id
                reused_project_ids.add(project.id)
            elif project_name:
                client_id = _parse_int(row.get("client_id"))
                client_name = str(row.get("client_name") or "").strip()
                project = None
                if client_id:
                    client = db.query(Client).filter(Client.id == client_id).first()
                    if not client:
                        raise ValueError(f"Client #{client_id} was not found")
                elif client_name:
                    client = _find_client_by_name(db, client_name)
                    if not client:
                        client = Client(
                            name=client_name,
                            preferred_currency=(str(row.get("currency") or "USD").strip().upper()[:3] or "USD"),
                            source="time_import",
                            owner_user_id=current_user.id,
                            created_at=now,
                            updated_at=now,
                        )
                        db.add(client)
                        db.flush()
                        result.created_clients += 1
                    client_id = client.id
                else:
                    project = _find_project_by_name(db, project_name)
                    if project:
                        client_id = project.client_id
                        reused_project_ids.add(project.id)
                    else:
                        currency = str(row.get("currency") or "USD").strip().upper()[:3] or "USD"
                        client, created_import_client = _get_or_create_import_client(db, current_user, now, currency)
                        if created_import_client:
                            result.created_clients += 1
                        client_id = client.id

                project = project or _find_project(db, project_name, client_id)
                if project:
                    reused_project_ids.add(project.id)
                else:
                    currency = str(row.get("currency") or "USD").strip().upper()[:3] or "USD"
                    project_hourly_rate = _parse_float(row.get("hourly_rate"))
                    project = Project(
                        client_id=client_id,
                        name=project_name,
                        billing_method=str(row.get("billing_method") or "hourly").strip() or "hourly",
                        hourly_rate=project_hourly_rate,
                        status="active",
                        currency=currency,
                        created_by=current_user.id,
                        created_at=now,
                        updated_at=now,
                    )
                    db.add(project)
                    db.flush()
                    created_project_ids.add(project.id)
            else:
                raise ValueError("Project name or project_id is required")

            task_id = None
            task = None
            task_name = str(row.get("task_name") or "").strip()
            if task_name:
                task = _find_task(db, project.id, task_name)
                if not task:
                    task = ProjectTask(
                        project_id=project.id,
                        name=task_name,
                        hourly_rate=_parse_float(row.get("hourly_rate")),
                        status="active",
                        created_at=now,
                        updated_at=now,
                    )
                    db.add(task)
                    db.flush()
                    created_task_ids.add(task.id)
                task_id = task.id

            hourly_rate = _resolve_hourly_rate(_parse_float(row.get("hourly_rate")), task, project)
            entry = TimeEntry(
                project_id=project.id,
                task_id=task_id,
                user_id=current_user.id,
                client_id=client_id,
                description=(str(row.get("description")).strip() if row.get("description") else None),
                notes=(str(row.get("notes")).strip() if row.get("notes") else None),
                started_at=started_at,
                ended_at=ended_at,
                duration_minutes=duration_minutes,
                hourly_rate=hourly_rate,
                billable=_parse_bool(row.get("billable"), default=True),
                status="logged",
                invoiced=False,
                created_at=now,
                updated_at=now,
            )
            entry.compute_amount()
            db.add(entry)
            result.created_time_entries += 1
        except Exception as exc:
            result.skipped_rows += 1
            result.errors.append(TimeImportError(row=index, message=str(exc)))

    if result.created_time_entries == 0:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail={"message": "No time entries could be imported", "errors": [e.model_dump() for e in result.errors]},
        )

    db.commit()
    result.created_projects = len(created_project_ids)
    result.reused_projects = len(reused_project_ids - created_project_ids)
    result.created_tasks = len(created_task_ids)

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="IMPORT",
        resource_type="time_entries",
        resource_id=file.filename,
        resource_name=f"Time CSV import: {file.filename}",
        details=result.model_dump(),
        status="success" if result.skipped_rows == 0 else "partial_success",
    )

    return result


# ---- Standard CRUD (AFTER static paths) ----

@time_entries_router.post("", response_model=TimeEntryResponse, status_code=201)
def create_time_entry(
    payload: TimeEntryCreate,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    """Manually log a time entry (not a timer start)."""
    project = db.query(Project).filter(Project.id == payload.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    now = datetime.now(timezone.utc)
    task = db.query(ProjectTask).filter(ProjectTask.id == payload.task_id).first() if payload.task_id else None
    hourly_rate = _resolve_hourly_rate(payload.hourly_rate if payload.hourly_rate > 0 else None, task, project)

    # Compute duration if ended_at is provided and duration_minutes is not
    duration_minutes = payload.duration_minutes
    ended_at = payload.ended_at
    if ended_at and duration_minutes is None:
        delta = ended_at - payload.started_at
        duration_minutes = max(1, int(delta.total_seconds() / 60))

    entry = TimeEntry(
        project_id=payload.project_id,
        task_id=payload.task_id,
        user_id=current_user.id,
        client_id=project.client_id,
        description=payload.description,
        notes=payload.notes,
        started_at=payload.started_at,
        ended_at=ended_at,
        duration_minutes=duration_minutes,
        hourly_rate=hourly_rate,
        billable=payload.billable,
        status="logged",
        invoiced=False,
        created_at=now,
        updated_at=now,
    )
    entry.compute_amount()
    db.add(entry)
    db.commit()
    db.refresh(entry)

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="CREATE",
        resource_type="time_entry",
        resource_id=str(entry.id),
        resource_name=f"Time Log for Project {payload.project_id}",
        details=payload.model_dump(),
        status="success"
    )

    return _enrich_time_entry(entry, db)


@time_entries_router.get("", response_model=List[TimeEntryResponse])
def list_time_entries(
    project_id: Optional[int] = Query(None),
    task_id: Optional[int] = Query(None),
    user_id: Optional[int] = Query(None),
    client_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    invoiced: Optional[bool] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    q = db.query(TimeEntry)
    if project_id:
        q = q.filter(TimeEntry.project_id == project_id)
    if task_id:
        q = q.filter(TimeEntry.task_id == task_id)
    if user_id:
        q = q.filter(TimeEntry.user_id == user_id)
    if client_id:
        q = q.filter(TimeEntry.client_id == client_id)
    if status:
        q = q.filter(TimeEntry.status == status)
    if invoiced is not None:
        q = q.filter(TimeEntry.invoiced == invoiced)
    entries = q.order_by(TimeEntry.started_at.desc()).offset(skip).limit(limit).all()
    return [_enrich_time_entry(e, db) for e in entries]


@time_entries_router.patch("/{entry_id}", response_model=TimeEntryResponse)
def update_time_entry(
    entry_id: int,
    payload: TimeEntryUpdate,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    entry = db.query(TimeEntry).filter(TimeEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Time entry not found")

    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(entry, field, value)

    # Recompute duration from started_at/ended_at if both are set and duration_minutes not explicit
    if entry.started_at and entry.ended_at and "duration_minutes" not in payload.model_dump(exclude_unset=True):
        delta = entry.ended_at - entry.started_at
        entry.duration_minutes = max(1, int(delta.total_seconds() / 60))

    entry.compute_amount()
    entry.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(entry)

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="UPDATE",
        resource_type="time_entry",
        resource_id=str(entry.id),
        resource_name=f"Time Log for Project {entry.project_id}",
        details=payload.model_dump(exclude_unset=True),
        status="success"
    )

    return _enrich_time_entry(entry, db)


@time_entries_router.delete("/{entry_id}", status_code=204)
def delete_time_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: MasterUser = Depends(get_current_user),
):
    entry = db.query(TimeEntry).filter(TimeEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Time entry not found")
    db.delete(entry)
    db.commit()

    log_audit_event(
        db=db,
        user_id=current_user.id,
        user_email=current_user.email,
        action="DELETE",
        resource_type="time_entry",
        resource_id=str(entry_id),
        resource_name=f"Time Log for Project {entry.project_id}",
        status="success"
    )
