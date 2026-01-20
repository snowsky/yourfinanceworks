"""
Report Export Handlers

This module provides export functionality for reports in multiple formats:
- PDF with professional formatting and company branding
- CSV for data analysis compatibility  
- Excel with multi-sheet support and formatting
"""

import csv
import io
import logging
from datetime import datetime
from typing import Dict, Any, List, Optional, Union
from pathlib import Path

# PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Local imports
from core.schemas.report import ReportData, ExportFormat

logger = logging.getLogger(__name__)


class ExportError(Exception):
    """Custom exception for export-related errors"""
    pass


class BaseExporter:
    """Base class for all export handlers"""
    
    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def export_report(self, report_data: ReportData, **kwargs) -> Union[bytes, str]:
        """Export report data to the specific format"""
        raise NotImplementedError("Subclasses must implement export_report method")
    
    def validate_report_data(self, report_data: ReportData) -> None:
        """Validate report data before export"""
        if not report_data:
            raise ExportError("Report data is required")
        
        if not report_data.data:
            raise ExportError("Report data cannot be empty")
        
        if not report_data.summary:
            raise ExportError("Report summary is required")


class PDFExporter(BaseExporter):
    """Export reports to PDF format with professional formatting and company branding"""
    
    def __init__(self, company_data: Optional[Dict[str, Any]] = None):
        super().__init__()
        self.company_data = company_data or {}
        self.styles = getSampleStyleSheet()
        self._create_custom_styles()
    
    def _create_custom_styles(self):
        """Create custom paragraph styles for professional formatting"""
        # Title style
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Heading1'],
            fontSize=20,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2c3e50')
        ))
        
        # Company name style
        self.styles.add(ParagraphStyle(
            name='CompanyName',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            alignment=TA_LEFT,
            textColor=colors.HexColor('#34495e')
        ))
        
        # Section header style
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading3'],
            fontSize=12,
            spaceBefore=15,
            spaceAfter=8,
            textColor=colors.HexColor('#2c3e50'),
            fontName='Helvetica-Bold'
        ))
        
        # Summary style
        self.styles.add(ParagraphStyle(
            name='Summary',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=15,
            textColor=colors.HexColor('#7f8c8d')
        ))
        
        # Table cell style for wrapping
        self.styles.add(ParagraphStyle(
            name='TableCell',
            parent=self.styles['Normal'],
            fontSize=9,
            leading=11,
            alignment=TA_LEFT,
            wordWrap='CJK'  # Better wrapping for various character sets
        ))
        
        self.styles.add(ParagraphStyle(
            name='TableHeader',
            parent=self.styles['Normal'],
            fontSize=10,
            leading=12,
            alignment=TA_CENTER,
            textColor=colors.whitesmoke,
            fontName='Helvetica-Bold'
        ))
    
    def export_report(self, report_data: ReportData, **kwargs) -> bytes:
        """Export report data to PDF format"""
        try:
            self.validate_report_data(report_data)
            
            # Create PDF buffer
            buffer = io.BytesIO()
            
            # Create document with professional margins
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=72
            )
            
            # Build document content
            story = []
            
            # Header with company branding
            story.extend(self._build_header(report_data))
            story.append(Spacer(1, 20))
            
            # Report summary
            story.extend(self._build_summary(report_data))
            story.append(Spacer(1, 20))
            
            # Report data table
            story.extend(self._build_data_table(report_data))
            
            # Footer
            story.extend(self._build_footer())
            
            # Build PDF
            doc.build(story)
            
            # Get PDF bytes
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            self.logger.info(f"Generated PDF report for {report_data.report_type}")
            return pdf_bytes
            
        except Exception as e:
            self.logger.error(f"Failed to generate PDF report: {str(e)}")
            raise ExportError(f"PDF export failed: {str(e)}")
    
    def _build_header(self, report_data: ReportData) -> List:
        """Build report header with company branding"""
        elements = []
        
        # Company name if available
        if self.company_data.get('name'):
            elements.append(Paragraph(self.company_data['name'], self.styles['CompanyName']))
        
        # Report title
        title = f"{report_data.report_type.value.title()} Report"
        elements.append(Paragraph(title, self.styles['ReportTitle']))
        
        # Generation date
        gen_date = report_data.metadata.generated_at.strftime('%B %d, %Y at %I:%M %p')
        elements.append(Paragraph(f"Generated on {gen_date}", self.styles['Summary']))
        
        return elements
    
    def _build_summary(self, report_data: ReportData) -> List:
        """Build report summary section"""
        elements = []
        
        elements.append(Paragraph("Report Summary", self.styles['SectionHeader']))
        
        # Summary data
        summary_data = [
            ['Total Records:', str(report_data.summary.total_records)],
            ['Date Range:', self._format_date_range(report_data.summary.date_range)],
        ]
        
        # Add total amount if available
        if report_data.summary.total_amount is not None:
            currency = report_data.summary.currency or 'USD'
            amount_str = f"{report_data.summary.total_amount:,.2f} {currency}"
            summary_data.append(['Total Amount:', amount_str])
        
        # Add key metrics
        for key, value in report_data.summary.key_metrics.items():
            if isinstance(value, (int, float)):
                if isinstance(value, float):
                    value_str = f"{value:,.2f}"
                else:
                    value_str = f"{value:,}"
            else:
                value_str = str(value)
            summary_data.append([f"{key.replace('_', ' ').title()}:", value_str])
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(summary_table)
        
        return elements
    
    def _build_data_table(self, report_data: ReportData) -> List:
        """Build main data table with intelligent column width handling"""
        elements = []

        if not report_data.data:
            elements.append(Paragraph("No data available", self.styles['Normal']))
            return elements

        elements.append(Paragraph("Report Data", self.styles['SectionHeader']))

        # Get column headers from first row
        first_row = report_data.data[0]
        headers = list(first_row.keys())
        num_cols = len(headers)

        # Build table data with improved text formatting
        table_data = [headers]  # Header row

        for row in report_data.data:
            table_row = []
            for header in headers:
                value = row.get(header, '')
                # Format values appropriately
                formatted_value = self._format_table_value(value, header)
                # Wrap in Paragraph for automatic wrapping/sizing in table cell
                p = Paragraph(str(formatted_value), self.styles['TableCell'])
                table_row.append(p)
            table_data.append(table_row)

        # Wrap headers in Paragraphs too for consistency and wrapping
        table_data[0] = [Paragraph(h, self.styles['TableHeader']) for h in headers]

        # Handle extreme cases with too many columns
        if num_cols > 25:
            # For extremely wide tables, create a simplified view or split into multiple tables
            elements.extend(self._build_wide_table_fallback(report_data, headers, table_data))
            return elements

        # Calculate intelligent column widths based on content
        col_widths = self._calculate_column_widths(headers, table_data, num_cols)

        # Adjust font sizes based on number of columns
        header_font_size, data_font_size = self._get_font_sizes(num_cols)

        # Create table with improved styling
        data_table = Table(table_data, colWidths=col_widths)
        data_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 6),

            # Data rows styling
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), data_font_size),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Changed to TOP for better multiline appearance
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))

        elements.append(data_table)

        return elements

    def _format_table_value(self, value: Any, header: str) -> str:
        """Format table values with intelligent text handling"""
        if value is None or value == '':
            return ''

        # Format numbers
        if isinstance(value, (int, float)):
            if isinstance(value, float):
                # Use shorter format for wide tables
                return f"{value:.2f}"
            else:
                return f"{value:,}"

        # Format dates
        elif isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')

        # Format dates
        elif isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')

        # Return as string (Paragraph will handle the wrapping)
        else:
            return str(value).replace('\n', '<br/>')

    def _calculate_column_widths(self, headers: List[str], table_data: List[List], num_cols: int) -> List[float]:
        """Calculate intelligent column widths based on content and available space"""
        available_width = 6.5 * inch  # Available width for table
        min_col_width = 0.4 * inch   # Minimum column width
        max_col_width = 1.5 * inch   # Maximum column width

        # For very wide tables, reduce minimum width
        if num_cols > 15:
            min_col_width = 0.3 * inch
        elif num_cols > 10:
            min_col_width = 0.35 * inch

        # Analyze content to determine optimal widths
        col_content_lengths = []
        for col_idx in range(num_cols):
            max_length = len(headers[col_idx])  # Start with header length

            # Check content lengths in data rows (sample first 10 rows for performance)
            for row_idx in range(1, min(len(table_data), 11)):
                if col_idx < len(table_data[row_idx]):
                    content = str(table_data[row_idx][col_idx])
                    # Count lines for multi-line content
                    lines = content.split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    max_length = max(max_length, max_line_length)

            col_content_lengths.append(max_length)

        # Calculate base widths based on content
        base_widths = []
        for length in col_content_lengths:
            # Estimate width based on character count (rough approximation)
            estimated_width = min(max(length * 0.08 * inch, min_col_width), max_col_width)
            base_widths.append(estimated_width)

        # Normalize widths to fit available space
        total_base_width = sum(base_widths)

        if total_base_width > available_width:
            # Scale down proportionally
            scale_factor = available_width / total_base_width
            col_widths = [width * scale_factor for width in base_widths]

            # Ensure minimum widths
            col_widths = [max(width, min_col_width) for width in col_widths]

            # If we still exceed available width after minimums, reduce proportionally again
            total_min_width = sum(col_widths)
            if total_min_width > available_width:
                scale_factor = available_width / total_min_width
                col_widths = [width * scale_factor for width in col_widths]
        else:
            # Distribute extra space proportionally
            extra_space = available_width - total_base_width
            if num_cols > 0:
                extra_per_col = extra_space / num_cols
                col_widths = [width + extra_per_col for width in base_widths]
                # Cap at maximum width
                col_widths = [min(width, max_col_width) for width in col_widths]

        return col_widths

    def _get_font_sizes(self, num_cols: int) -> tuple[int, int]:
        """Get appropriate font sizes based on number of columns"""
        if num_cols > 20:
            return (8, 7)  # Very small for very wide tables
        elif num_cols > 15:
            return (9, 8)  # Small for wide tables
        elif num_cols > 10:
            return (10, 9)  # Medium-small for moderately wide tables
        else:
            return (10, 9)  # Standard sizes for narrow tables

    def _build_wide_table_fallback(self, report_data: ReportData, headers: List[str], table_data: List[List]) -> List:
        """Handle extremely wide tables by creating a simplified summary or multiple tables"""
        elements = []

        # Add a warning about the wide table
        elements.append(Paragraph(
            "Note: This report contains many columns. Showing summarized view below.",
            self.styles['Summary']
        ))
        elements.append(Spacer(1, 10))

        # Create a summary table with key columns only
        key_columns = self._identify_key_columns(headers, table_data)

        if key_columns:
            # Build summary table with just the most important columns
            summary_headers = [headers[i] for i in key_columns]
            summary_data = [summary_headers]  # Header row

            for row in table_data[1:]:  # Skip header row
                summary_row = [row[i] for i in key_columns if i < len(row)]
                if summary_row:
                    summary_data.append(summary_row)

            # Create summary table
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))

            elements.append(Paragraph("Summary View (Key Columns Only)", self.styles['SectionHeader']))
            elements.append(summary_table)
        else:
            # Fallback to showing just row counts
            total_rows = len(table_data) - 1  # Subtract header row
            elements.append(Paragraph(
                f"This report contains {len(headers)} columns and {total_rows} rows. " +
                "For detailed data, please use CSV or Excel export formats.",
                self.styles['Normal']
            ))

        return elements

    def _identify_key_columns(self, headers: List[str], table_data: List[List], max_cols: int = 8) -> List[int]:
        """Identify the most important columns to show in summary view"""
        key_column_names = [
            'id', 'name', 'title', 'date', 'amount', 'total', 'status',
            'created_at', 'updated_at', 'client', 'customer'
        ]

        key_indices = []
        remaining_slots = max_cols

        # First, prioritize key columns
        for i, header in enumerate(headers):
            header_lower = header.lower()
            if any(key_name in header_lower for key_name in key_column_names):
                key_indices.append(i)
                remaining_slots -= 1
                if remaining_slots <= 0:
                    break

        # If we still have slots, add the first few remaining columns
        if remaining_slots > 0:
            for i, header in enumerate(headers):
                if i not in key_indices:
                    key_indices.append(i)
                    remaining_slots -= 1
                    if remaining_slots <= 0:
                        break

        return sorted(key_indices)

    def _build_footer(self) -> List:
        """Build report footer"""
        elements = []
        
        elements.append(Spacer(1, 30))
        
        footer_text = "This report was generated automatically."
        if self.company_data.get('name'):
            footer_text = f"Generated by {self.company_data['name']} reporting system."
        
        elements.append(Paragraph(footer_text, self.styles['Summary']))
        
        return elements
    
    def _format_date_range(self, date_range: Optional[Dict[str, datetime]]) -> str:
        """Format date range for display"""
        if not date_range:
            return "All time"
        
        start_date = date_range.get('start')
        end_date = date_range.get('end')
        
        if start_date and end_date:
            return f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        elif start_date:
            return f"From {start_date.strftime('%Y-%m-%d')}"
        elif end_date:
            return f"Until {end_date.strftime('%Y-%m-%d')}"
        else:
            return "All time"


class CSVExporter(BaseExporter):
    """Export reports to CSV format for data analysis compatibility"""
    
    def export_report(self, report_data: ReportData, **kwargs) -> str:
        """Export report data to CSV format"""
        try:
            self.validate_report_data(report_data)
            
            # Create CSV buffer
            output = io.StringIO()
            
            if not report_data.data:
                return ""
            
            # Get column headers from first row
            first_row = report_data.data[0]
            headers = list(first_row.keys())
            
            # Create CSV writer
            writer = csv.DictWriter(output, fieldnames=headers)
            
            # Write header
            writer.writeheader()
            
            # Write data rows
            for row in report_data.data:
                # Convert values to strings and handle None values
                csv_row = {}
                for header in headers:
                    value = row.get(header)
                    if value is None:
                        csv_row[header] = ''
                    elif isinstance(value, datetime):
                        csv_row[header] = value.isoformat()
                    elif isinstance(value, (int, float)):
                        csv_row[header] = str(value)
                    else:
                        csv_row[header] = str(value)
                
                writer.writerow(csv_row)
            
            csv_content = output.getvalue()
            output.close()
            
            self.logger.info(f"Generated CSV report for {report_data.report_type}")
            return csv_content
            
        except Exception as e:
            self.logger.error(f"Failed to generate CSV report: {str(e)}")
            raise ExportError(f"CSV export failed: {str(e)}")


class ExcelExporter(BaseExporter):
    """Export reports to Excel format with multi-sheet support and formatting"""
    
    def __init__(self, company_data: Optional[Dict[str, Any]] = None):
        super().__init__()
        self.company_data = company_data or {}
    
    def export_report(self, report_data: ReportData, **kwargs) -> bytes:
        """Export report data to Excel format"""
        try:
            self.validate_report_data(report_data)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create main data sheet
            ws_data = wb.create_sheet(title="Report Data")
            self._build_data_sheet(ws_data, report_data)
            
            # Create summary sheet
            ws_summary = wb.create_sheet(title="Summary")
            self._build_summary_sheet(ws_summary, report_data)
            
            # Set active sheet to summary
            wb.active = ws_summary
            
            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()
            
            self.logger.info(f"Generated Excel report for {report_data.report_type}")
            return excel_bytes
            
        except Exception as e:
            self.logger.error(f"Failed to generate Excel report: {str(e)}")
            raise ExportError(f"Excel export failed: {str(e)}")
    
    def _build_summary_sheet(self, ws: Worksheet, report_data: ReportData):
        """Build summary sheet with report overview"""
        # Title
        ws['A1'] = f"{report_data.report_type.value.title()} Report Summary"
        ws['A1'].font = Font(size=16, bold=True, color='2C3E50')
        ws.merge_cells('A1:B1')
        
        # Company name if available
        row = 3
        if self.company_data.get('name'):
            ws[f'A{row}'] = "Company:"
            ws[f'B{row}'] = self.company_data['name']
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Generation date
        ws[f'A{row}'] = "Generated:"
        ws[f'B{row}'] = report_data.metadata.generated_at.strftime('%Y-%m-%d %H:%M:%S')
        ws[f'A{row}'].font = Font(bold=True)
        row += 2
        
        # Summary data
        ws[f'A{row}'] = "Total Records:"
        ws[f'B{row}'] = report_data.summary.total_records
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Date range
        ws[f'A{row}'] = "Date Range:"
        ws[f'B{row}'] = self._format_date_range_excel(report_data.summary.date_range)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Total amount if available
        if report_data.summary.total_amount is not None:
            ws[f'A{row}'] = "Total Amount:"
            currency = report_data.summary.currency or 'USD'
            ws[f'B{row}'] = f"{report_data.summary.total_amount:,.2f} {currency}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Key metrics
        if report_data.summary.key_metrics:
            row += 1
            ws[f'A{row}'] = "Key Metrics:"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for key, value in report_data.summary.key_metrics.items():
                ws[f'A{row}'] = f"{key.replace('_', ' ').title()}:"
                if isinstance(value, (int, float)):
                    if isinstance(value, float):
                        ws[f'B{row}'] = f"{value:,.2f}"
                    else:
                        ws[f'B{row}'] = f"{value:,}"
                else:
                    ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _build_data_sheet(self, ws: Worksheet, report_data: ReportData):
        """Build data sheet with report data"""
        if not report_data.data:
            ws['A1'] = "No data available"
            return
        
        # Get column headers from first row
        first_row = report_data.data[0]
        headers = list(first_row.keys())
        
        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data rows
        for row_idx, row_data in enumerate(report_data.data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                
                # Format values appropriately
                if value is None:
                    cell_value = ''
                elif isinstance(value, datetime):
                    # openpyxl does not support timezone-aware datetimes
                    cell_value = value.replace(tzinfo=None) if value.tzinfo else value
                elif isinstance(value, (int, float)):
                    cell_value = value
                else:
                    cell_value = str(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Apply alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        
        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(report_data.data) + 1, 
                               min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _format_date_range_excel(self, date_range: Optional[Dict[str, datetime]]) -> str:
        """Format date range for Excel display"""
        if not date_range:
            return "All time"
        
        start_date = date_range.get('start')
        end_date = date_range.get('end')
        
        if start_date and end_date:
            return f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        elif start_date:
            return f"From {start_date.strftime('%Y-%m-%d')}"
        elif end_date:
            return f"Until {end_date.strftime('%Y-%m-%d')}"
        else:
            return "All time"


class ReportExportService:
    """Main service for handling report exports in multiple formats"""
    
    def __init__(self, company_data: Optional[Dict[str, Any]] = None):
        self.company_data = company_data or {}
        self.logger = logging.getLogger(__name__)
        
        # Initialize exporters
        self.exporters = {
            ExportFormat.PDF: PDFExporter(company_data),
            ExportFormat.CSV: CSVExporter(),
            ExportFormat.EXCEL: ExcelExporter(company_data),
        }
    
    def export_report(
        self, 
        report_data: ReportData, 
        export_format: ExportFormat,
        **kwargs
    ) -> Union[bytes, str]:
        """Export report in the specified format"""
        try:
            # Validate export format
            if export_format not in self.exporters:
                raise ExportError(f"Unsupported export format: {export_format}")
            
            # Get appropriate exporter
            exporter = self.exporters[export_format]
            
            # Export report
            result = exporter.export_report(report_data, **kwargs)
            
            self.logger.info(f"Successfully exported {report_data.report_type} report as {export_format}")
            return result
            
        except Exception as e:
            self.logger.error(f"Export failed for format {export_format}: {str(e)}")
            raise ExportError(f"Export failed: {str(e)}")
    
    def get_supported_formats(self) -> List[ExportFormat]:
        """Get list of supported export formats"""
        return list(self.exporters.keys())
    
    def validate_export_format(self, export_format: str) -> ExportFormat:
        """Validate and convert export format string to enum"""
        try:
            return ExportFormat(export_format.lower())
        except ValueError:
            supported = [fmt.value for fmt in self.get_supported_formats()]
            raise ExportError(f"Invalid export format '{export_format}'. Supported formats: {supported}")